import pandas as pd
from openpyxl.styles import PatternFill

def generate_excel_diff_report(differences: list, output_path: str):
    """Génère un export xlsx propre à partir des différences, avec code couleur rouge."""
    if not differences:
        pd.DataFrame([{"Message": "Aucune différence"}]).to_excel(output_path, index=False)
        return

    df = pd.DataFrame(differences)
    # Assurons l'ordre des colonnes si possible
    if "Valeur 1" in df.columns and "Valeur 2" in df.columns:
        # Ordre par défaut si non fourni
        pass  # df est déjà structuré par le main si besoin

    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Différences')
    
    worksheet = writer.sheets['Différences']
    
    # Fond rouge pastel pour les décalages ("Valeur 1" et "Valeur 2")
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Application du code couleur aux deux dernières colonnes itérativement
    num_cols = len(df.columns)
    val1_idx = None
    val2_idx = None
    
    # Trouver l'index des colonnes de valeurs (généralement les deux dernières)
    for idx, col_name in enumerate(df.columns, 1):
        if col_name not in ["Contexte", "Identifiant", "Champ"]:
            if val1_idx is None:
                val1_idx = idx
            else:
                val2_idx = idx

    if val1_idx and val2_idx:
        for row in range(2, len(df) + 2):
            cell1 = worksheet.cell(row=row, column=val1_idx)
            cell2 = worksheet.cell(row=row, column=val2_idx)
            
            # Formater en rouge si cellules différentes (normalement toutes dans ce df)
            if cell1.value != cell2.value:
                cell1.fill = red_fill
                cell2.fill = red_fill
             
    # Ajuster la largeur des colonnes
    for column_cells in worksheet.columns:
        length = max((len(str(cell.value) or "") for cell in column_cells), default=10)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 60)
        
    writer.close()
